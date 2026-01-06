from openpyxl import load_workbook

# Use ORIGINAL file (not updated) to check formulas and colors
file_path = "Interarch Building Solutions Ltd l Valuation template l Dec 25 (1).xlsx"
wb = load_workbook(file_path)
ws = wb['Factsheet']

print("FACTSHEET ANALYSIS - Finding Blue cells and Formula cells")
print("=" * 80)

formula_rows = set()
blue_rows = set()
black_data_rows = set()

for row in range(1, 61):
    label = ws.cell(row=row, column=1).value
    
    if label and str(label).strip():
        has_formula = False
        has_blue = False
        has_black_data = False
        
        for col in range(3, 9):  # Columns C to H
            cell = ws.cell(row=row, column=col)
            
            # Check for formula
            if cell.value and str(cell.value).startswith('='):
                has_formula = True
            
            # Check for blue color
            if cell.font and cell.font.color:
                color = cell.font.color
                try:
                    rgb = str(color.rgb) if color.rgb else ""
                    # Common blue colors in Excel
                    if any(blue in rgb for blue in ['0070C0', '00B0F0', '4472C4', '5B9BD5', '2E75B6']):
                        has_blue = True
                    elif rgb in ['FF000000', '00000000'] or color.theme in [0, 1]:
                        if cell.value is not None:
                            has_black_data = True
                except:
                    if cell.value is not None:
                        has_black_data = True
            else:
                if cell.value is not None:
                    has_black_data = True
        
        label_str = str(label).replace('\n', ' ')[:45]
        
        if has_formula:
            formula_rows.add(row)
            print(f"Row {row:2d} [FORMULA]: {label_str}")
        elif has_blue:
            blue_rows.add(row)
            print(f"Row {row:2d} [BLUE   ]: {label_str}")
        elif has_black_data:
            black_data_rows.add(row)
            print(f"Row {row:2d} [BLACK  ]: {label_str}")

print("\n" + "=" * 80)
print("SUMMARY")
print("=" * 80)
print(f"\nFORMULA rows (DO NOT TOUCH): {sorted(formula_rows)}")
print(f"BLUE rows (DO NOT TOUCH): {sorted(blue_rows)}")
print(f"BLACK data rows (FILL FROM BASE DATA): {sorted(black_data_rows)}")

print("\n\nROWS TO UPDATE IN AUTOMATION SCRIPT:")
print("-" * 40)
for row in sorted(black_data_rows):
    label = ws.cell(row=row, column=1).value
    label_str = str(label).replace('\n', ' ')[:45] if label else ""
    print(f"  Row {row}: {label_str}")
