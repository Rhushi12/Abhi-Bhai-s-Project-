import pandas as pd
from openpyxl import load_workbook
import logging
import sys
import os
import re
import json

# Try to import tkinter for file dialog
try:
    import tkinter as tk
    from tkinter import filedialog
    HAS_TKINTER = True
except ImportError:
    HAS_TKINTER = False

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

def normalize_name(text):
    """Normalize field name for matching: lowercase, remove special chars."""
    if pd.isna(text) or text is None:
        return ""
    return re.sub(r'[^a-zA-Z0-9]', '', str(text)).lower()

def clean_value(val):
    """Convert value to float, handling NaN and string formatting."""
    if pd.isna(val):
        return 0.0
    try:
        return float(str(val).replace(',', '').strip())
    except ValueError:
        return 0.0

def select_file():
    """Open a file dialog to select an Excel file. Uses native Mac dialog on macOS."""
    import platform
    import subprocess
    
    # On Mac, use native AppleScript file dialog (more reliable than tkinter)
    if platform.system() == 'Darwin':
        try:
            script = '''
            tell application "System Events"
                activate
                set theFile to choose file with prompt "Select Excel File" of type {"xlsx", "xls"}
                return POSIX path of theFile
            end tell
            '''
            result = subprocess.run(['osascript', '-e', script], capture_output=True, text=True)
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout.strip()
            else:
                print("File selection cancelled or failed.")
                return None
        except Exception as e:
            print(f"Error with file dialog: {e}")
            return None
    
    # On Windows/Linux, use tkinter
    if not HAS_TKINTER:
        print("ERROR: tkinter not available. Please provide file path as argument.")
        return None
    
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    )
    
    root.destroy()
    return file_path if file_path else None

def is_blue_cell(cell):
    """Check if a cell has blue font color."""
    if cell.font and cell.font.color:
        color = cell.font.color
        try:
            rgb = str(color.rgb) if color.rgb else ""
            blue_colors = ['0070C0', '00B0F0', '4472C4', '5B9BD5', '2E75B6', '1F4E79', '305496']
            if any(blue in rgb for blue in blue_colors):
                return True
        except:
            pass
    return False

def has_formula(cell):
    """Check if a cell contains a formula."""
    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
        return True
    return False

# =========================================================
# FIELD NAME MAPPINGS
# =========================================================
# Maps Factsheet field names -> Base data field name variations
# Format: 'factsheet_normalized_name': ['basedata_variation1', 'basedata_variation2', ...]
# =========================================================

FIELD_MAPPINGS = {
    # P&L Data mappings
    'sales': ['sales', 'salesrevenue', 'revenue', 'netsales'],
    'costofsales': ['expenses', 'costofsales', 'costofgoodssold', 'cogs'],
    'operatingincome': ['operatingprofit', 'operatingincome'],
    'otherincomenormal': ['otherincomenormal', 'otherincome'],
    'otherincomeexceptional': ['exceptionalitems', 'exceptional'],
    'otherexpensesexceptionalitems': [],  # No source - default 0
    'depreciation': ['depreciation', 'depreciationamortization'],
    'interest': ['interest', 'interestexpense', 'financecharges'],
    'tax': ['taxinrcrores'],  # Specific match to avoid 'profit before tax'
    'equitydividend': ['dividendpayout', 'dividend'],  # May default to 0
    
    # Balance Sheet mappings
    'networth': ['networth', 'shareholdersequity', 'totalequity'],
    'longtermborrowings': ['longtermborrowings', 'longtermborrowing', 'longtermdebt'],
    'shorttermborrowings': ['shorttermborrowings', 'shorttermborrowing', 'shorttermdebt'],
    'otherborrowingsleasereserves': ['leaseliabilities', 'leasereserves', 'otherborrowings'],
    'debt': ['borrowings', 'totaldebt', 'debt'],
    'fixedassets': ['fixedassets', 'grossblock', 'propertyplantequipment'],
    'capitalwipassets': ['cwip', 'capitalworkinprogress', 'capitalwip'],
    'currentassest': ['otherassets', 'currentassets', 'totalcurrentassets'],
    'currentliabilities': ['otherliabilities', 'currentliabilities', 'totalcurrentliabilities'],
    'investments': ['investments', 'longtermandinvestments'],
    'cashcashequivalent': ['cashequivalents', 'cashandbank', 'cash'],
    'currentinvestmentsmarketablesecurities': [],  # No source - default 0
    'debtors': ['tradereceivables', 'debtors', 'accountsreceivable'],
    'inventories': ['inventories', 'inventory', 'stock'],
    'reservesandsurplusesbalance': ['reserves', 'reservesandsurplus'],
    
    # Share transaction fields
    'amountreceivedonissuanceofnewshares': ['amountreceivedonissuanceofnewshares', 'issuanceofnewshares', 'shareissuance'],
    'amountspendforbuybackofshares': ['amountspendforbuybackofshares', 'buybackofshares', 'sharebuyback'],
}

# Load overrides from JSON file if available (for dynamic editing)
# Helper to reload mappings
def reload_mappings():
    """Reload mappings from JSON file to ensure latest config is used."""
    try:
        json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'mappings.json')
        if os.path.exists(json_path):
            with open(json_path, 'r') as f:
                custom_mappings = json.load(f)
                FIELD_MAPPINGS.update(custom_mappings)
                logger.info("✓ Reloaded custom field mappings")
    except Exception as e:
        logger.warning(f"⚠ Could not reload mappings.json: {e}")

# Fields that should default to 0 if not found in Base data
DEFAULT_ZERO_FIELDS = [
    'otherexpensesexceptionalitems',
    'currentinvestmentsmarketablesecurities',
    'equitydividend',
]

# Fields that should show 'input' placeholder (require manual entry)
INPUT_PLACEHOLDER_FIELDS = [
    'amountreceivedonissuanceofnewshares',
    'amountspendforbuybackofshares',
]


def find_row_by_name(df, field_variations):
    """Find the row index in dataframe that matches any of the field name variations.
    Prioritizes exact matches over partial matches to avoid false positives."""
    
    # First pass: look for exact matches only
    for idx, row in df.iterrows():
        label = row.iloc[0]  # First column contains labels
        normalized_label = normalize_name(label)
        
        for variation in field_variations:
            if normalized_label == variation:
                return idx
    
    # Second pass: look for partial matches (variation is contained in label)
    # Only if no exact match was found
    for idx, row in df.iterrows():
        label = row.iloc[0]
        normalized_label = normalize_name(label)
        
        for variation in field_variations:
            # Partial match: label starts with variation (more restrictive)
            if normalized_label.startswith(variation):
                return idx
    
    return None


def extract_base_data(base_df):
    """Extract all data from Base data sheet using field name matching."""
    data = {}
    matched_fields = []
    
    for factsheet_field, base_variations in FIELD_MAPPINGS.items():
        if not base_variations:
            # No source in base data - check which default to use
            if factsheet_field in INPUT_PLACEHOLDER_FIELDS:
                data[factsheet_field] = ['input'] * 6
            else:
                data[factsheet_field] = [0.0] * 6
            continue
        
        row_idx = find_row_by_name(base_df, base_variations)
        
        if row_idx is not None:
            try:
                values = [clean_value(base_df.iloc[row_idx, col]) for col in range(1, 7)]
                data[factsheet_field] = values
                matched_fields.append(f"✓ {factsheet_field} → row {row_idx}")
            except Exception as e:
                # Use appropriate default on error
                if factsheet_field in INPUT_PLACEHOLDER_FIELDS:
                    data[factsheet_field] = ['input'] * 6
                else:
                    data[factsheet_field] = [0.0] * 6
                matched_fields.append(f"⚠ {factsheet_field} → error: {e}")
        else:
            # Not found - use appropriate default
            if factsheet_field in INPUT_PLACEHOLDER_FIELDS:
                data[factsheet_field] = ['input'] * 6
                matched_fields.append(f"⚠ {factsheet_field} → NOT FOUND (using 'input')")
            else:
                data[factsheet_field] = [0.0] * 6
                if factsheet_field not in DEFAULT_ZERO_FIELDS:
                    matched_fields.append(f"⚠ {factsheet_field} → NOT FOUND (using 0)")
    
    return data, matched_fields


def find_factsheet_rows(ws, max_row=60):
    """Build a map of normalized field names to their row numbers in Factsheet."""
    row_map = {}
    for row in range(1, max_row + 1):
        cell = ws.cell(row=row, column=1)
        label = cell.value
        if label:
            normalized = normalize_name(label)
            if normalized and normalized not in row_map:
                row_map[normalized] = row
    return row_map


def automate_transfer(file_path, output_path=None):
    """
    Automate data transfer from Base data sheet to Factsheet.
    Uses field name matching for robustness (not row positions).
    Only writes to BLACK cells (skips BLUE cells and cells with formulas).
    """
    
    if output_path is None:
        base_name = os.path.splitext(file_path)[0]
        output_path = f"{base_name}_updated.xlsx"
    
    print("=" * 60)
    print("BASE DATA TO FACTSHEET AUTOMATION")
    print("(Using Field Name Matching)")
    print("=" * 60)
    logger.info(f"\nInput file: {file_path}")
    logger.info(f"Output file: {output_path}\n")
    
    # Reload mappings to ensure we use the latest config
    reload_mappings()
    
    # Read Base data sheet
    try:
        base_df = pd.read_excel(file_path, sheet_name='Base data', header=None)
        logger.info(f"✓ Base data loaded: {base_df.shape[0]} rows, {base_df.shape[1]} columns")
    except Exception as e:
        print(f"\n❌ ERROR: Could not read 'Base data' sheet: {e}")
        return None
    
    # Extract data using field name matching
    print("\nMatching fields from Base data...")
    print("-" * 40)
    data, match_log = extract_base_data(base_df)
    for log_entry in match_log:
        logger.info(f"  {log_entry}")
    logger.info(f"\n✓ Extracted {len(data)} fields")
    
    # Load workbook
    print("\nAnalyzing Factsheet structure...")
    try:
        wb = load_workbook(file_path)
        ws = wb['Factsheet']
    except Exception as e:
        print(f"\n❌ ERROR: Could not open 'Factsheet' sheet: {e}")
        return None
    
    # Build Factsheet row map
    factsheet_row_map = find_factsheet_rows(ws)
    logger.info(f"✓ Found {len(factsheet_row_map)} labeled rows in Factsheet")
    
    # Write values to Factsheet
    print("\nWriting to Factsheet (skipping Blue and Formula cells)...")
    print("-" * 40)
    cells_written = 0
    cells_skipped = 0
    fields_not_found = []
    
    for factsheet_field, values in data.items():
        # Find the row in Factsheet for this field
        if factsheet_field in factsheet_row_map:
            excel_row = factsheet_row_map[factsheet_field]
        else:
            # Try partial match
            found = False
            for fs_name, fs_row in factsheet_row_map.items():
                if factsheet_field in fs_name or fs_name in factsheet_field:
                    excel_row = fs_row
                    found = True
                    break
            if not found:
                fields_not_found.append(factsheet_field)
                continue
        
        # Write values to columns C-H
        for col_offset, val in enumerate(values):
            col = 3 + col_offset  # Start from column C
            cell = ws.cell(row=excel_row, column=col)
            
            # Check if cell should be skipped
            if is_blue_cell(cell):
                cells_skipped += 1
                continue
            
            if has_formula(cell):
                cells_skipped += 1
                continue
            
            # Write value to black cell
            cell.value = val
            cells_written += 1
    
    logger.info(f"\n✓ Written {cells_written} cells")
    logger.info(f"✓ Skipped {cells_skipped} cells (Blue or Formula)")
    
    if fields_not_found:
        logger.info(f"\n⚠ Fields not found in Factsheet: {fields_not_found}")
    
    # Save the workbook
    try:
        wb.save(output_path)
        print("\n" + "=" * 60)
        print(f"✅ SUCCESS! Output saved to:")
        print(f"   {output_path}")
        print("=" * 60)
    except Exception as e:
        print(f"\n❌ ERROR saving file: {e}")
        print("Make sure the output file is not open in Excel")
        return None
    
    return output_path


def main():
    """Main entry point."""
    print("\n" + "=" * 60)
    print("     VALUATION TEMPLATE AUTOMATION")
    print("     Base Data → Factsheet (Field Name Matching)")
    print("=" * 60 + "\n")
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        if not os.path.exists(file_path):
            print(f"❌ File not found: {file_path}")
            input("\nPress Enter to exit...")
            return
    else:
        print("Opening file selector...")
        file_path = select_file()
        
        if not file_path:
            print("❌ No file selected. Exiting.")
            input("\nPress Enter to exit...")
            return
    
    print(f"Selected: {os.path.basename(file_path)}\n")
    
    result = automate_transfer(file_path)
    
    if result:
        print("\n💡 TIP: Script uses field name matching - works even if rows move!")
    
    input("\nPress Enter to exit...")


if __name__ == "__main__":
    main()